#!/usr/bin/env python3
"""One-time / regen: Excel → schedule-raw.json (UTF-8). Merge syllabus in schedule.json manually or via editor."""
import json
import os
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise SystemExit('pip install openpyxl')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS = os.path.join(os.path.dirname(ROOT), 'assets', 'summerschool')
OUT = os.path.join(ROOT, 'data', 'schedule-raw.json')


def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%m/%d')
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(v))
    if m:
        return f'{int(m.group(2))}/{int(m.group(3))}'
    return str(v).strip()


def cell_text(v):
    if v is None:
        return ''
    return str(v).strip().replace('\n', ' / ')


def main():
    xlsx_files = [f for f in os.listdir(ASSETS) if f.endswith('.xlsx')]
    if not xlsx_files:
        raise SystemExit(f'No .xlsx in {ASSETS}')
    wb = openpyxl.load_workbook(os.path.join(ASSETS, xlsx_files[0]), data_only=True)
    ws = wb.active

    online = []
    for r in range(3, 9):
        online.append({
            'date': fmt_date(ws.cell(r, 3).value),
            'time': cell_text(ws.cell(r, 4).value),
            'title': cell_text(ws.cell(r, 5).value),
            'content': cell_text(ws.cell(r, 6).value),
        })

    day_headers = []
    for c in range(2, 13):
        day_headers.append({
            'weekday': cell_text(ws.cell(9, c).value),
            'date': fmt_date(ws.cell(10, c).value),
        })

    rows = []
    for r in range(11, ws.max_row + 1):
        cells = [cell_text(ws.cell(r, c).value) for c in range(2, 13)]
        rows.append({'time': cell_text(ws.cell(r, 1).value), 'cells': cells})

    out = {
        'title': cell_text(ws.cell(1, 1).value),
        'online': online,
        'offline': {'dayHeaders': day_headers, 'rows': rows},
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print('Wrote', OUT)


if __name__ == '__main__':
    main()
